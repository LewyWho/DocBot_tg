import asyncio
import os
import sqlite3
import time
import uuid
import zipfile
from zipfile import ZipFile

import openpyxl
from aiogram.utils.exceptions import BadRequest
from aiogram import types, Dispatcher, Bot
from aiogram.contrib.middlewares.logging import LoggingMiddleware
from aiogram.contrib.fsm_storage.memory import MemoryStorage
from aiogram.dispatcher import FSMContext
from aiogram.dispatcher.filters import state
from aiogram.dispatcher.filters.state import StatesGroup, State
from aiogram.types import ParseMode, InlineKeyboardButton, InlineKeyboardMarkup
from aiogram.utils import executor
import pytz
from openpyxl.styles import Alignment, Font

import config
import keyboards
import register
import sms
from datetime import datetime
from aiogram.dispatcher import FSMContext

bot = Bot(token=config.BOT_TOKEN, parse_mode=ParseMode.HTML)
dp = Dispatcher(bot, storage=MemoryStorage())
dp.middleware.setup(LoggingMiddleware())
conn = sqlite3.connect('database.db')
cursor = conn.cursor()


class AddResult(StatesGroup):
    event_id = State()
    result = State()
    points = State()


class CreateEvent(StatesGroup):
    name = State()
    description = State()
    level = State()
    file_id = State()
    date = State()


class AddNewStaff(StatesGroup):
    staff_user_id = State()
    staff_full_name = State()
    staff_department = State()
    staff_role = State()
    staff_additional_info = State()


################################ START ################################
@dp.message_handler(commands=['start'])
async def handler_start(message: types.Message, state: FSMContext):
    verify = cursor.execute("SELECT * FROM users WHERE user_id = ?", (message.from_user.id,)).fetchone()
    if verify:
        await state.finish()
        await bot.send_message(chat_id=message.from_user.id,
                               text=sms.my_profile_in_main_menu(message.from_user.id),
                               reply_markup=await keyboards.main_keyboard(message.from_user.id))


# ROLE 1 CREATE MP
@dp.callback_query_handler(text='create_event')
async def create_event(call: types.CallbackQuery):
    await call.message.answer('Введите название мероприятия:')
    await CreateEvent.name.set()


@dp.message_handler(state=CreateEvent.name)
async def set_description(message: types.Message, state: FSMContext):
    async with state.proxy() as data:
        data['name'] = message.text
    await message.answer('Введите описание мероприятия:')
    await CreateEvent.next()


@dp.message_handler(state=CreateEvent.description)
async def set_file_id(message: types.Message, state: FSMContext):
    async with state.proxy() as data:
        data['description'] = message.text
    keyboard = types.InlineKeyboardMarkup(row_width=2)
    buttons = [
        types.InlineKeyboardButton(text="Городской", callback_data="level_1"),
        types.InlineKeyboardButton(text="Конкурсное мероприятие", callback_data="level_2"),
        types.InlineKeyboardButton(text="Международный", callback_data="level_3"),
        types.InlineKeyboardButton(text="Окружной", callback_data="level_4"),
        types.InlineKeyboardButton(text="Районный", callback_data="level_5"),
        types.InlineKeyboardButton(text="Региональный", callback_data="level_6"),
        types.InlineKeyboardButton(text="Учрежденческий", callback_data="level_7"),
        types.InlineKeyboardButton(text="Федеральный", callback_data="level_8"),
    ]
    buttons.sort(key=lambda x: x.text)
    keyboard.add(*buttons)
    await message.answer("Выберите уровень мероприятия:", reply_markup=keyboard)
    await CreateEvent.next()


@dp.callback_query_handler(lambda call: call.data.startswith('level_'), state=CreateEvent.level)
async def set_level(call: types.CallbackQuery, state: FSMContext):
    async with state.proxy() as data:
        keyboard_button_text = next(
            button.text for row in call.message.reply_markup.inline_keyboard for button in row if
            button.callback_data == call.data)
        data['level'] = keyboard_button_text
    await call.message.answer('Отправьте файл мероприятия (если есть, если нет, то напишите "Нет"):')
    await CreateEvent.next()


@dp.message_handler(state=CreateEvent.file_id, content_types=[
    types.ContentType.DOCUMENT,
    types.ContentType.PHOTO,
    types.ContentType.TEXT
])
async def set_date(message: types.Message, state: FSMContext):
    async with state.proxy() as data:
        if message.content_type == types.ContentType.TEXT:
            data['file_id'] = message.text
        elif message.document:
            file_id = message.document.file_id

            file_extension = message.document.file_name.split('.')[-1]
            file_path = os.path.join('docs', f'{file_id}.{file_extension}')

            data['file_id'] = f"{file_id}.{file_extension}"

            await message.document.download(destination=file_path)
        elif message.photo:
            file_id = message.photo[-1].file_id

            unique_filename = str(uuid.uuid4())

            photo_info = await bot.get_file(file_id)

            file_extension = photo_info.file_path.split('.')[-1]

            file_path = os.path.join('docs', f'{unique_filename}.{file_extension}')

            await bot.download_file(photo_info.file_path, destination=file_path)

            data['file_id'] = f"{unique_filename}.{file_extension}"

    await message.answer('Введите дату мероприятия в формате ДД-ММ-ГГГГ:')
    await CreateEvent.next()


@dp.message_handler(state=CreateEvent.date)
async def save_event(message: types.Message, state: FSMContext):
    async with state.proxy() as data:
        try:
            event_date = datetime.strptime(message.text, '%d-%m-%Y')
            data['event_date'] = event_date.strftime('%d-%m-%Y')
        except ValueError:
            await message.answer('Неверный формат даты. Попробуйте снова.')
            return

        cursor.execute('''
            INSERT INTO events (event_name, event_description, event_level, file_id, event_date, responsible_user_id)
            VALUES (?, ?, ?, ?, ?, ?)
        ''', (
            data['name'], data['description'], data['level'], data['file_id'], data['event_date'],
            message.from_user.id))
        conn.commit()
    await message.answer('Мероприятие сохранено!')
    await state.finish()


# ROLE 1 check_my_stats
@dp.callback_query_handler(text='check_my_stats')
async def check_my_stats(call: types.CallbackQuery):
    user_id = call.from_user.id

    events_query = """
        SELECT events.event_id, events.event_name, events.event_description, events.event_level, events.file_id, events.event_date, users.full_name
        FROM events
        JOIN users ON events.responsible_user_id = users.user_id
        WHERE events.responsible_user_id = ?
    """
    events = cursor.execute(events_query, (user_id,)).fetchall()

    if not events:
        await call.message.answer("У вас нет мероприятий.")
        return

    await call.message.answer('Данные загружаются, пожалуйста, подождите...')

    zip_filename = f"events_{user_id}.zip"
    filename = f"events_{user_id}.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Мероприятия"
    bold_font = openpyxl.styles.Font(bold=True)
    ws.append(
        ["ID", "Название мероприятия", "Описание мероприятия", "Уровень мероприятия", "Файл", "Дата мероприятия",
         "Ответственный за мероприятие"])
    for cell in ws[1]:
        cell.font = bold_font

    for event in events:
        event_list = list(event)
        if event_list[4] and (event_list[4].lower() != 'нет'):
            file_path = f'files/{event_list[4]}'
            event_list[4] = f'=HYPERLINK("{file_path}", "Есть файлы (Кликабельно)")'
        ws.append(event_list)

    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center', wrap_text=True)

    results_query = """
        SELECT events.event_id, events.event_name, events.event_description, events.event_level, results.result, results.points
        FROM results
        JOIN events ON results.event_id = events.event_id
        WHERE events.responsible_user_id = ?
    """
    results = cursor.execute(results_query, (user_id,)).fetchall()

    # Создание листа "Результаты мероприятий"
    ws_results = wb.create_sheet(title="Результаты мероприятий")

    # Добавление заголовков столбцов
    ws_results.append(
        ["ID события", "Название мероприятия", "Описание мероприятия", "Уровень мероприятия", "Результат", "Баллы"])

    # Заполнение данными
    for result in results:
        ws_results.append(result)

    # Настройка форматирования
    for cell in ws_results[1]:
        cell.font = bold_font

    for row in ws_results.iter_rows():
        for cell in row:
            cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center', wrap_text=True)

    wb.save(filename)

    with zipfile.ZipFile(zip_filename, 'w') as zipf:
        zipf.writestr('Мои мероприятия/', '')

        zipf.write(filename, os.path.join('Мои мероприятия', filename))

        zipf.writestr(os.path.join('Мои мероприятия', 'files/'), '')

        docs_folder = 'docs'
        for root, dirs, files in os.walk(docs_folder):
            for file in files:
                file_id, _ = os.path.splitext(file)
                file_path = os.path.join(root, file)
                zipf.write(file_path,
                           arcname=os.path.join('Мои мероприятия', 'files', os.path.relpath(file_path, docs_folder)))

    with open(zip_filename, "rb") as file:
        await call.message.answer_document(file)

    os.remove(filename)
    os.remove(zip_filename)


# ROLE 1 add_result
@dp.callback_query_handler(text='add_result')
async def add_result(call: types.CallbackQuery):
    user_id = call.from_user.id

    # Запрос для выбора мероприятий без результата
    not_results_query = """
            SELECT events.event_id, events.event_name, events.event_description, events.event_level, events.file_id, events.event_date, users.full_name
            FROM events
            JOIN users ON events.responsible_user_id = users.user_id
            LEFT JOIN results ON events.event_id = results.event_id
            WHERE events.responsible_user_id = ? AND results.result IS NULL
        """
    not_results_data = cursor.execute(not_results_query, (user_id,)).fetchall()

    if not not_results_data:
        await call.message.answer("Все ваши мероприятия имеют результат.")
        return

    await call.message.answer('Вот ваши мероприятия, где не указан результат:')

    for event in not_results_data:
        event_info = f"ID Мероприятия: {event[0]}\nНазвание: {event[1]}\nОписание: {event[2]}\nУровень: {event[3]}\nДата: {event[5]}\nОтветственный: {event[6]}\n➖➖➖➖➖➖➖➖➖➖➖➖➖➖\n"
        await call.message.answer(event_info)
    await call.message.answer('Для добавления результата введите ID мероприятия:')
    await AddResult.event_id.set()


@dp.message_handler(state=AddResult.event_id)
async def save_result(message: types.Message, state: FSMContext):
    async with state.proxy() as data:
        try:
            event_id = int(message.text)
            data['event_id'] = event_id
        except ValueError:
            await message.answer('Неверный формат ID. Попробуйте снова.')
            return

        keyboard = types.InlineKeyboardMarkup(row_width=3)
        buttons = [
            types.InlineKeyboardButton(text="1 место", callback_data="result_1"),
            types.InlineKeyboardButton(text="2 место", callback_data="result_2"),
            types.InlineKeyboardButton(text="3 место", callback_data="result_3"),
            types.InlineKeyboardButton(text="Участие", callback_data="result_4"),
        ]
        buttons.sort(key=lambda x: x.text)
        keyboard.add(*buttons)

        await message.answer('Укажите ваш результат:', reply_markup=keyboard)
        await AddResult.next()


@dp.callback_query_handler(lambda call: call.data.startswith('result_'), state=AddResult.result)
async def save_result(call: types.CallbackQuery, state: FSMContext):
    async with state.proxy() as data:
        keyboard_button_text = next(
            button.text for row in call.message.reply_markup.inline_keyboard for button in row if
            button.callback_data == call.data)
        data['result'] = keyboard_button_text

        result = data['result']

    if result == '1 место':
        points = 10
    elif result == '2 место':
        points = 7
    elif result == '3 место':
        points = 5
    elif result == 'Участие':
        points = 1
    else:
        points = 0

    event_level = cursor.execute("SELECT event_level FROM events WHERE event_id =?", (data['event_id'],)).fetchone()[0]

    if event_level == 'Городской':
        points = points * 1
    elif event_level == 'Конкурсное мероприятие':
        points = points * 2
    elif event_level == 'Международный':
        points = points * 5
    elif event_level == 'Окружной':
        points = points * 1.5
    elif event_level == 'Районный':
        points = points * 1
    elif event_level == 'Региональный':
        points = points * 2
    elif event_level == 'Учрежденческий':
        points = points * 1
    elif event_level == 'Федеральный':
        points = points * 3

    cursor.execute("INSERT INTO results (event_id, result, points) VALUES (?,?,?)",
                   (data['event_id'], result, points))

    conn.commit()

    await call.message.answer("Результат успешно добавлен!")
    await state.finish()


# ROLE 2 check_stats_staff
@dp.callback_query_handler(text='check_stats_staff')
async def check_stats_staff(call: types.CallbackQuery):
    user_id = call.from_user.id

    staff_query = """
        SELECT users.full_name, users.department, events.event_name, events.event_description, events.event_level, events.file_id, events.event_date, results.result, results.points
        FROM users
        LEFT JOIN events ON users.user_id = events.responsible_user_id
        LEFT JOIN results ON events.event_id = results.event_id
        WHERE users.role = 1
    """
    staff_data = cursor.execute(staff_query).fetchall()

    if not staff_data:
        await call.message.answer("Нет данных о сотрудниках.")
        return

    await call.message.answer('Данные загружаются, пожалуйста, подождите...')

    zip_filename = f"staff_stats.zip"

    wb = openpyxl.Workbook()

    bold_font = openpyxl.styles.Font(bold=True)

    # Создаем словарь для хранения данных по отделам
    department_data = {}

    # Группируем данные по отделам
    for staff_row in staff_data:
        full_name, department, *event_data = staff_row
        if department not in department_data:
            department_data[department] = []
        department_data[department].append([full_name, *event_data])

    # Создаем листы для каждого отдела
    for department, data in department_data.items():
        ws_staff = wb.create_sheet(title=department)
        ws_staff.append(
            ["ФИО сотрудника", "Название мероприятия", "Описание мероприятия", "Уровень мероприятия", "Файл",
             "Дата мероприятия", "Результат", "Баллы"])
        for staff_row in data:
            if staff_row[4] and (staff_row[4].lower() != 'нет'):
                file_path = f'files/{staff_row[4]}'
                staff_row[4] = f'=HYPERLINK("{file_path}", "Есть файлы (Кликабельно)")'
            ws_staff.append(staff_row)

        for cell in ws_staff[1]:
            cell.font = bold_font

        for row in ws_staff.iter_rows():
            for cell in row:
                cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center', wrap_text=True)

    # Удаляем лист по умолчанию
    default_sheet = wb.get_sheet_by_name('Sheet')
    wb.remove(default_sheet)

    # Сохраняем файл Excel и создаем архив
    filename = f"staff_stats.xlsx"
    wb.save(filename)

    with zipfile.ZipFile(zip_filename, 'w') as zipf:
        zipf.write(filename, os.path.join('Сотрудники', filename))

        docs_folder = 'docs'
        for root, dirs, files in os.walk(docs_folder):
            for file in files:
                file_id, _ = os.path.splitext(file)
                file_path = os.path.join(root, file)
                zipf.write(file_path,
                           arcname=os.path.join('Сотрудники', 'files', os.path.relpath(file_path, docs_folder)))

    with open(zip_filename, "rb") as file:
        await call.message.answer_document(file)

    os.remove(filename)
    os.remove(zip_filename)


# ROLE 4 CREATE USER
@dp.callback_query_handler(text='add_new_staff')
async def add_new_staff(call: types.CallbackQuery):
    await call.answer("Введите user_id сотрудника: ")
    await AddNewStaff.staff_user_id.set()


@dp.message_handler(state=AddNewStaff.staff_user_id)
async def add_staff_full_name(message: types.Message, state: FSMContext):
    await state.update_data(staff_user_id=message.text)
    await message.answer("Введите полное имя сотрудника: ")
    await AddNewStaff.staff_full_name.set()


@dp.message_handler(state=AddNewStaff.staff_full_name)
async def add_staff_department(message: types.Message, state: FSMContext):
    async with state.proxy() as data:
        data['staff_full_name'] = message.text
    await message.answer("Введите отдел сотрудника: ")
    await AddNewStaff.staff_department.set()


@dp.message_handler(state=AddNewStaff.staff_department)
async def add_staff_role(message: types.Message, state: FSMContext):
    async with state.proxy() as data:
        data['staff_department'] = message.text
    await message.answer("Введите должность сотрудника (1 - Сотрудник, 2 - Администрация НГИЭУ, 3 - Аналитик, "
                         "4 - Администратор бота): ")
    await AddNewStaff.staff_role.set()


@dp.message_handler(state=AddNewStaff.staff_role)
async def add_staff_additional_info(message: types.Message, state: FSMContext):
    async with state.proxy() as data:
        data['staff_role'] = message.text
    await message.answer("Введите дополнительную информацию о сотруднике: ")
    await AddNewStaff.staff_additional_info.set()


@dp.message_handler(state=AddNewStaff.staff_additional_info)
async def save_staff_info(message: types.Message, state: FSMContext):
    async with state.proxy() as data:
        data['staff_additional_info'] = message.text
        cursor.execute(
            "INSERT INTO users (user_id, full_name, department, role, additional_info) VALUES (?, ?, ?, ?, ?)",
            (data['staff_user_id'], data['staff_full_name'], data['staff_department'], data['staff_role'],
             data['staff_additional_info']))
        conn.commit()
        await message.answer("Данные о сотруднике сохранены.")
    await state.finish()


class DelStaff(StatesGroup):
    staff_user_id = State()


@dp.callback_query_handler(text='del_staff')
async def del_staff_info(call: types.CallbackQuery):
    await call.answer("Введите user_id сотрудника, которого нужно удалить: ")
    await DelStaff.staff_user_id.set()


@dp.message_handler(state=DelStaff.staff_user_id)
async def del_staff(message: types.Message, state: FSMContext):
    async with state.proxy() as data:
        data['staff_user_id'] = message.text
        cursor.execute("DELETE FROM users WHERE user_id = ?", (data['staff_user_id'],))
        conn.commit()
        await message.answer("Данные о сотруднике удалены.")
    await state.finish()


class ChangeStaff(StatesGroup):
    staff_info_to_change = State()
    staff_user_id = State()
    staff_change = State()


@dp.callback_query_handler(text='change_staff')
async def change_staff_info(call: types.CallbackQuery):
    await call.message.answer("Введите user_id сотрудника, которого нужно изменить: ")
    await ChangeStaff.staff_user_id.set()


@dp.message_handler(state=ChangeStaff.staff_user_id)
async def change_staff(message: types.Message, state: FSMContext):
    async with state.proxy() as data:
        data['staff_user_id'] = message.text
        cursor.execute("SELECT * FROM users WHERE user_id = ?", (data['staff_user_id'],))
        staff_info = cursor.fetchone()
        if staff_info:
            keyboard = types.InlineKeyboardMarkup(row_width=2)
            buttons = [
                types.InlineKeyboardButton(text="user_id сотрудника", callback_data="change_0"),
                types.InlineKeyboardButton(text="ФИО сотрудника", callback_data="change_1"),
                types.InlineKeyboardButton(text="Отдел сотрудника", callback_data="change_2"),
                types.InlineKeyboardButton(text="Должность сотрудника", callback_data="change_3"),
                types.InlineKeyboardButton(text="Доп. инф. сотрудника", callback_data="change_4")
            ]
            buttons.sort(key=lambda x: x.text)
            keyboard.add(*buttons)

            await message.answer(f"user_id сотрудника: {staff_info[0]}\n"
                                 f"Полное имя сотрудника: {staff_info[1]}\n"
                                 f"Отдел сотрудника: {staff_info[2]}\n"
                                 f"Должность сотрудника: {staff_info[3]}\n"
                                 f"Дополнительная информация о сотруднике: {staff_info[4]}\n➖➖➖➖➖➖➖➖➖➖➖➖➖➖\nЧто вы хотите изменить?",
                                 reply_markup=keyboard)
            await ChangeStaff.staff_info_to_change.set()


@dp.callback_query_handler(lambda call: call.data.startswith('change_'), state=ChangeStaff.staff_info_to_change)
async def change_staff_info_select(call: types.CallbackQuery, state: FSMContext):
    async with state.proxy() as data:
        keyboard_button_text = next(
            button.text for row in call.message.reply_markup.inline_keyboard for button in row if
            button.callback_data == call.data)
        data['changes'] = keyboard_button_text

    await call.message.answer("Введите новое значение:")

    await ChangeStaff.staff_change.set()


@dp.message_handler(state=ChangeStaff.staff_change)
async def change_staff_info_value(message: types.Message, state: FSMContext):
    async with state.proxy() as data:
        colum_name = data['changes']
        print(colum_name)
    if colum_name == 'user_id сотрудника':
        staff_user_id = message.text
        cursor.execute("UPDATE users SET user_id =? WHERE user_id =?", (message.text, data['staff_user_id']))
    elif colum_name == 'ФИО сотрудника':
        staff_full_name = message.text
        cursor.execute("UPDATE users SET full_name =? WHERE user_id =?", (message.text, data['staff_user_id']))
    elif colum_name == 'Отдел сотрудника':
        staff_department = message.text
        cursor.execute("UPDATE users SET department =? WHERE user_id =?", (message.text, data['staff_user_id']))
    elif colum_name == 'Должность сотрудника':
        staff_role = message.text
        cursor.execute("UPDATE users SET role =? WHERE user_id =?", (message.text, data['staff_user_id']))
    elif colum_name == 'Доп. инф. сотрудника':
        staff_additional_info = message.text
        cursor.execute("UPDATE users SET additional_info =? WHERE user_id =?", (message.text, data['staff_user_id']))

    conn.commit()

    await message.answer("Данные изменены.")
    await state.finish()



if __name__ == '__main__':
    dp.register_message_handler(handler_start, commands=['start'], state='*')
    register.all_callback(dp)
    executor.start_polling(dp, skip_updates=True)
